"""
UTM Command Dashboard — v10  (utm_v10_FINAL)
FastAPI Backend

FILE:  main.py
ROLE:  The entire Python backend lives here.
       It loads simulation data, enriches it with derived columns,
       exposes 13 REST API endpoints consumed by the frontend,
       and serves the static frontend files.

ENTRY POINT:
    python main.py         → starts Uvicorn on 0.0.0.0:8000

DATA FLOW:
    1. POST /api/connect/postgres  or  POST /api/connect/excel
       → load_postgres() / load_excel() normalises column names
         and stores two DataFrames in the global _S dict.
    2. GET /api/trial/{tid}/{endpoint}?path_runs=1,2
       → _get() filters _S by trial + path runs, calls enrich()
         to compute derived columns (bat_u, cx, cy, da, eff, …),
         then the matching data-builder function assembles the
         JSON response.

DEFAULT DB:  utm_geo_db  (localhost:5432, user=postgres)
"""

from __future__ import annotations
import io, json, warnings
from pathlib import Path
from typing   import Optional

import numpy  as np
import pandas as pd
import uvicorn
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

warnings.filterwarnings("ignore")

BASE     = Path(__file__).parent
FRONTEND = BASE / "frontend"

app = FastAPI(title="DRONEVIEW v2")
app.mount("/static", StaticFiles(directory=str(FRONTEND)), name="static")
app.mount("/js",  StaticFiles(directory=str(FRONTEND / "js")),  name="js")
app.mount("/css", StaticFiles(directory=str(FRONTEND / "css")), name="css")
app.mount("/images", StaticFiles(directory=str(FRONTEND)), name="images")

# ── Global in-memory session ─────────────────────────────────────────────────
# _S holds all loaded data in RAM.  No database round-trip on every request.
# Restarting the server clears _S and requires re-connecting.
# Keys:
#   "drones"     → pd.DataFrame  one row per drone flight record
#   "collisions" → pd.DataFrame  one row per collision event
#   "source"     → str           "postgresql" or "excel"
#   "trials"     → list[str]     unique trial_id values
_S: dict = {
    "drones":     pd.DataFrame(),
    "collisions": pd.DataFrame(),
    "source":     "none",
    "trials":     [],
}

# ── Serialiser ─────────────────────────────────────────────────────────────────
# J(obj)     → wraps any Python dict/list in a JSONResponse,
#               first stripping NaN/Inf values so JSON stays valid.
# _clean(o)  → recursive cleaner called by J().
# _s(v)      → used inside data builders to clean individual values
#               (numpy int/float types, NaN/Inf → None).
def J(obj): return JSONResponse(content=_clean(obj))
def _clean(o):
    if isinstance(o, dict):  return {k: _clean(v) for k,v in o.items()}
    if isinstance(o, list):  return [_clean(v) for v in o]
    if isinstance(o, (np.integer,)):  return int(o)
    if isinstance(o, (np.floating,)):
        if np.isnan(o) or np.isinf(o): return None
        return float(o)
    if isinstance(o, float):
        if np.isnan(o) or np.isinf(o): return None
    return o
def _s(v):
    if isinstance(v, float) and (np.isnan(v) or np.isinf(v)): return None
    if isinstance(v, (np.integer,)): return int(v)
    if isinstance(v, (np.floating,)): return float(v)
    return v

# ── Parsers ────────────────────────────────────────────────────────────────────
# parse_battery(v)  → strips "%" suffix, returns float or NaN.
#                     Handles "85", "85%", "85.3 %" etc.
# parse_speed(v)    → strips "m/s" suffix, returns float or NaN.
# layer_to_alt(l)   → maps layer number → altitude metres:
#                     1→0m, 2→50m, 3→100m, 4→150m
def parse_battery(v):
    try:   return float(str(v).replace('%','').strip())
    except: return np.nan

def parse_speed(v):
    try:   return float(str(v).replace('m/s','').strip())
    except: return np.nan

def layer_to_alt(l):
    try:   return {1:0,2:50,3:100,4:150}.get(int(l), 0)
    except: return 0

# ── PostgreSQL loader ──────────────────────────────────────────────────────────
# load_postgres()  → connects via SQLAlchemy, reads two tables:
#     drone_summary_geo  JOIN simulation_runs_geo  → df_d (drones)
#     collision_log_geo  JOIN simulation_runs_geo  → df_c (collisions)
# Column names from the DB are renamed to the internal schema
# (e.g. "final_lat" → "coord_x", "vehicle_type" → "vehicle").
# Returns: {"drones": df_d, "collisions": df_c, "source": "postgresql"}
def load_postgres(host, port, dbname, user, password) -> dict:
    from sqlalchemy import create_engine
    eng = create_engine(
        f"postgresql+psycopg2://{user}:{password}@{host}:{port}/{dbname}",
        connect_args={"connect_timeout": 10})

    df_d = pd.read_sql(
        "SELECT d.*, r.run_label AS trial_id, r.trial_number "
        "FROM drone_summary_geo d "
        "JOIN simulation_runs_geo r USING (run_id)", eng)

    df_c = pd.read_sql(
        "SELECT c.*, r.run_label AS trial_id "
        "FROM collision_log_geo c "
        "JOIN simulation_runs_geo r USING (run_id)", eng)

    # Normalise column names → internal schema
    d_map = {
        'vehicle_type':       'vehicle',
        'battery_start':      'battery_start',
        'battery_end':        'battery_end',
        'battery_used':       'battery_consumed',
        'speed':              'drone_speed',
        'final_lat':          'coord_x',
        'final_lon':          'coord_y',
        'layer':              'drone_layer',
        'layer_name':         'layer_name',
        'distance_planned_m': 'distance_planned',
        'distance_actual_m':  'distance_actual',
    }
    df_d = df_d.rename(columns={k:v for k,v in d_map.items() if k in df_d.columns})

    c_map = {
        'crash_lat':       'drone_a_coord_x',
        'crash_lon':       'drone_a_coord_y',
        'crash_layer':     'drone_a_layer',
        'drone_a_battery': 'drone_a_battery_start',
        'drone_b_battery': 'drone_b_battery_start',
        'drone_a_distance':'drone_a_distance_actual',
        'drone_b_distance':'drone_b_distance_actual',
    }
    df_c = df_c.rename(columns={k:v for k,v in c_map.items() if k in df_c.columns})
    for col in ['drone_b_coord_x','drone_b_coord_y','drone_b_layer']:
        if col not in df_c.columns: df_c[col] = np.nan

    return {"drones": df_d, "collisions": df_c, "source": "postgresql"}

# ── Excel loader ───────────────────────────────────────────────────────────────
# load_excel(buf)  → reads a simulation Excel workbook using openpyxl.
# Expected sheet names:
#   "Run 1 — <label>", "Run 2 — <label>", …   drone flight records
#   "Collision Log"                             collision events
# Row 1 of each sheet: ignored (title row).
# Row 2: column headers (exact names matter — see docs/DATA_FORMAT.md).
# Row 3+: data.  Rows with blank "Drone ID" / "Event ID" are skipped.
# All sheets are merged into two DataFrames using the same internal schema
# as load_postgres().
# Returns: {"drones": df_d, "collisions": df_c, "source": "excel"}
def load_excel(buf) -> dict:
    import openpyxl
    wb    = openpyxl.load_workbook(buf, data_only=True)
    sheets = wb.sheetnames
    run_sheets = [s for s in sheets if s.startswith("Run ")]
    coll_sheet = "Collision Log"

    import re
    drone_rows = []
    for sname in run_sheets:
        m = re.match(r'Run\s+(\d+)\s*[—\-]+\s*(.+)', sname)
        pr  = int(m.group(1)) if m else 1
        lbl = m.group(2).strip() if m else sname
        ws  = wb[sname]
        rows= list(ws.iter_rows(values_only=True))
        hdrs= [str(c).strip() if c else "" for c in rows[1]]
        for row in rows[2:]:
            if not row[0]: continue
            r = dict(zip(hdrs, row))
            drone_rows.append({
                "trial_id":        "excel_trial",
                "trial_number":    1,
                "path_run":        pr,
                "path_label":      lbl,
                "drone_id":        _safe_int(r.get("Drone ID")),
                "vehicle":         r.get("Vehicle Type"),
                "flight_status":   r.get("Flight Status"),
                "battery_start":   r.get("Battery Start (%)"),
                "battery_end":     r.get("Battery End (%)"),
                "battery_consumed":r.get("Battery Used (%)"),
                "drone_speed":     r.get("Speed"),
                "coord_x":         r.get("Final Latitude"),
                "coord_y":         r.get("Final Longitude"),
                "drone_layer":     r.get("Layer"),
                "layer_name":      r.get("Layer Name"),
                "distance_planned":0,
                "distance_actual": r.get("Distance Actual (m)"),
                "source_node":     r.get("Source Node"),
                "destination_node":r.get("Destination Node"),
                "start_time":      str(r.get("Start Time","") or ""),
                "end_time":        str(r.get("End Time","") or ""),
            })

    coll_rows = []
    if coll_sheet in wb.sheetnames:
        ws   = wb[coll_sheet]
        rows = list(ws.iter_rows(values_only=True))
        hdrs = [str(c).strip() if c else "" for c in rows[1]]
        for row in rows[2:]:
            if not row[0]: continue
            r = dict(zip(hdrs, row))
            coll_rows.append({
                "trial_id":            "excel_trial",
                "path_run":            _safe_int(r.get("Path Run")),
                "path_label":          r.get("Path Label"),
                "event_id":            _safe_int(r.get("Event ID")),
                "grid_tick":           _safe_int(r.get("Grid Tick")),
                "type":                r.get("Type"),
                "severity":            r.get("Severity"),
                "collision_type":      r.get("Collision Type"),
                "drone_a_id":          _safe_int(r.get("Drone A ID")),
                "drone_a_vehicle":     r.get("Drone A Vehicle"),
                "drone_b_id":          _safe_int(r.get("Drone B ID")),
                "drone_b_vehicle":     r.get("Drone B Vehicle"),
                "drone_a_coord_x":     r.get("Crash Latitude"),
                "drone_a_coord_y":     r.get("Crash Longitude"),
                "drone_a_layer":       _safe_int(r.get("Crash Layer")),
                "drone_b_coord_x":     None,
                "drone_b_coord_y":     None,
                "drone_b_layer":       None,
                "drone_a_battery_start": r.get("Drone A Battery"),
                "drone_b_battery_start": r.get("Drone B Battery"),
                "drone_a_distance_actual": None,
                "drone_b_distance_actual": None,
            })

    df_d = pd.DataFrame(drone_rows)
    df_c = pd.DataFrame(coll_rows)
    return {"drones": df_d, "collisions": df_c, "source": "excel"}

# _safe_int(v) → converts any value to int, returns None on failure.
# Used by load_excel() when parsing Drone ID / Event ID columns that
# may come through as float (e.g. 42.0) or string from openpyxl.
def _safe_int(v):
    try: return int(str(v).strip())
    except: return None

# ── Enrich ─────────────────────────────────────────────────────────────────────
# enrich(df_d, df_c) → takes a filtered pair of DataFrames and adds
# derived columns that all data-builder functions and page JS files use.
#
# Drone DataFrame (df_d) derived columns added:
#   bat_s   battery_start %   (parsed float, NaN-safe)
#   bat_e   battery_end %
#   bat_u   battery_consumed %
#   spd     speed m/s
#   cx      final latitude  (coord_x coerced to float)
#   cy      final longitude (coord_y coerced to float)
#   layer   altitude layer 1–4 (drone_layer coerced to float)
#   alt     metres (layer_to_alt)
#   dp      planned distance m
#   da      actual distance m
#   eff     dp/da route efficiency ratio
#   duration  seconds (end_dt − start_dt)
#
# Collision DataFrame (df_c) derived columns added:
#   gtn     grid_tick coerced to float
#   drone_a_coord_x/y, drone_b_coord_x/y, drone_a/b_layer  → numeric
#   bat_a, bat_b   battery of each drone at collision time
#   da_a, da_b     actual distance of each drone
def enrich(df_d: pd.DataFrame, df_c: pd.DataFrame):
    d = df_d.copy()
    c = df_c.copy()
    if not d.empty:
        d["bat_s"]  = d["battery_start"].apply(parse_battery)
        d["bat_e"]  = d["battery_end"].apply(parse_battery)
        d["bat_u"]  = d["battery_consumed"].apply(parse_battery)
        d["spd"]    = d.get("drone_speed", pd.Series(dtype=float)).apply(parse_speed)
        d["cx"]     = pd.to_numeric(d.get("coord_x"), errors="coerce")
        d["cy"]     = pd.to_numeric(d.get("coord_y"), errors="coerce")
        d["layer"]  = pd.to_numeric(d.get("drone_layer"), errors="coerce")
        d["alt"]    = d["layer"].apply(layer_to_alt)
        d["dp"]     = pd.to_numeric(d.get("distance_planned", pd.Series(dtype=float)), errors="coerce")
        d["da"]     = pd.to_numeric(d.get("distance_actual"),  errors="coerce")
        d["eff"]    = (d["dp"] / d["da"].replace(0, np.nan)).round(3)
        d["start_dt"] = pd.to_datetime(d.get("start_time"), errors="coerce")
        d["end_dt"]   = pd.to_datetime(d.get("end_time"),   errors="coerce")
        d["duration"] = (d["end_dt"] - d["start_dt"]).dt.total_seconds()
    if not c.empty:
        c["gtn"] = pd.to_numeric(c.get("grid_tick"), errors="coerce")
        for col in ["drone_a_coord_x","drone_a_coord_y",
                    "drone_b_coord_x","drone_b_coord_y",
                    "drone_a_layer",  "drone_b_layer"]:
            c[col] = pd.to_numeric(c.get(col, pd.Series(dtype=float)), errors="coerce")
        c["bat_a"] = c.get("drone_a_battery_start", pd.Series(dtype=object)).apply(parse_battery)
        c["bat_b"] = c.get("drone_b_battery_start", pd.Series(dtype=object)).apply(parse_battery)
        c["da_a"]  = pd.to_numeric(c.get("drone_a_distance_actual"), errors="coerce")
        c["da_b"]  = pd.to_numeric(c.get("drone_b_distance_actual"), errors="coerce")
    return d, c

# ── _get helper ────────────────────────────────────────────────────────────────
# _get(tid, path_runs) → called at the start of every API endpoint.
# 1. Raises HTTP 400 if no data is loaded (_S["drones"] is empty).
# 2. Filters _S["drones"] and _S["collisions"] to the given trial_id.
# 3. If path_runs is a comma-separated string (e.g. "1,2,3"), further
#    filters both DataFrames to those path run numbers.
# 4. Calls enrich() and returns (d, c) — enriched drone and collision DFs.
# Every GET /api/trial/{tid}/... endpoint starts with:
#     d, c = _get(tid, path_runs)
def _get(tid: str, path_runs: Optional[str] = None):
    if _S["drones"].empty:
        raise HTTPException(400, "No data loaded")
    d = _S["drones"][_S["drones"]["trial_id"] == tid].copy()
    c = (_S["collisions"][_S["collisions"]["trial_id"] == tid].copy()
         if not _S["collisions"].empty else pd.DataFrame())
    if path_runs:
        prs = [int(x) for x in path_runs.split(",") if x.strip().isdigit()]
        if prs:
            d = d[d["path_run"].isin(prs)]
            if not c.empty and "path_run" in c.columns:
                c = c[c["path_run"].isin(prs)]
    return enrich(d, c)

# ═══════════════════════════════════════════════════════════════════════════════
# DATA BUILDERS
# Each function receives enriched (d, c) DataFrames and returns a plain
# Python dict/list that becomes the JSON response body via J().
# They are pure functions — no side effects, no global state access.
# ═══════════════════════════════════════════════════════════════════════════════

# compute_kpis(d, c)
# → The core KPI dict shared by Overview and Fleet pages.
# Counts drones by flight_status prefix and collisions by type.
# Returns: N, n_ok, n_crash, n_batt, n_canc, n_coll, n_dir, n_prox, n_nm,
#          comp_pct, crash_pct, collision_free_pct, avg_bat_used, avg_efficiency, avg_da
def compute_kpis(d, c):
    N = len(d)
    if N == 0:
        return dict(N=0, n_ok=0, n_crash=0, n_batt=0, n_canc=0,
                    n_coll=0, n_dir=0, n_prox=0, n_nm=0,
                    comp_pct=0.0, crash_pct=0.0, collision_free_pct=100.0,
                    avg_bat_used=0.0, avg_efficiency=0.0, avg_da=0.0)
    n_ok    = int(d["flight_status"].str.startswith("Complete").sum())
    n_crash = int(d["flight_status"].str.startswith("Collision").sum())
    n_batt  = int(d["flight_status"].str.startswith("Incomplete").sum())
    n_canc  = int(d["flight_status"].str.startswith("Cancelled").sum())
    n_coll  = len(c)
    n_dir   = int((c["type"]=="Direct").sum())    if n_coll else 0
    n_prox  = int((c["type"]=="Proximity").sum()) if n_coll else 0
    n_nm    = int((c["type"]=="Near Miss").sum())  if n_coll else 0
    avg_da  = float(d["da"].mean()) if "da" in d.columns and not d["da"].isna().all() else 0.0
    return dict(
        N=N, n_ok=n_ok, n_crash=n_crash, n_batt=n_batt, n_canc=n_canc,
        n_coll=n_coll, n_dir=n_dir, n_prox=n_prox, n_nm=n_nm,
        comp_pct=round(n_ok/N*100,1), crash_pct=round(n_crash/N*100,1),
        collision_free_pct=round((N-n_crash)/N*100,1),
        avg_bat_used=round(float(d["bat_u"].mean()),1) if "bat_u" in d.columns else 0.0,
        avg_efficiency=round(float(d["eff"].median()),3) if "eff" in d.columns and not d["eff"].isna().all() else 0.0,
        avg_da=round(avg_da,1),
    )

# path_run_comparison(d, c)
# → One summary row per path run, used in the Overview page
#   for the multi-run comparison chart and table.
# Returns: list[{path_run, label, total, complete, crash, battery,
#                cancelled, comp_pct, crash_pct, n_dir, n_prox, n_nm,
#                bat_avg, da_avg, events}]
def path_run_comparison(d, c):
    if "path_run" not in d.columns: return []
    rows = []
    for pr, grp in d.groupby("path_run"):
        n = len(grp)
        if n == 0: continue
        ok    = int(grp["flight_status"].str.startswith("Complete").sum())
        crash = int(grp["flight_status"].str.startswith("Collision").sum())
        batt  = int(grp["flight_status"].str.startswith("Incomplete").sum())
        canc  = int(grp["flight_status"].str.startswith("Cancelled").sum())
        tc    = c[c["path_run"]==pr] if not c.empty and "path_run" in c.columns else pd.DataFrame()
        n_dir = int((tc["type"]=="Direct").sum())    if not tc.empty else 0
        n_prox= int((tc["type"]=="Proximity").sum()) if not tc.empty else 0
        n_nm  = int((tc["type"]=="Near Miss").sum())  if not tc.empty else 0
        bat   = round(float(grp["bat_u"].mean()),1) if "bat_u" in grp.columns and not grp["bat_u"].isna().all() else 0
        da    = round(float(grp["da"].mean()),1) if "da" in grp.columns and not grp["da"].isna().all() else 0
        label = str(grp["path_label"].iloc[0]) if "path_label" in grp.columns else f"Run {pr}"
        rows.append(dict(
            path_run=int(pr), label=label, total=n,
            complete=ok, crash=crash, battery=batt, cancelled=canc,
            comp_pct=round(ok/n*100,1), crash_pct=round(crash/n*100,1),
            n_dir=n_dir, n_prox=n_prox, n_nm=n_nm,
            bat_avg=bat, da_avg=da, events=len(tc),
        ))
    return sorted(rows, key=lambda x: x["path_run"])

# vehicle_performance(d, c)
# → One row per vehicle type, used by Safety and Fleet pages.
# Returns: list[{vehicle, total, complete, crash, battery, cancelled,
#                comp_pct, crash_pct, bat_avg, bat_reserve,
#                avg_distance, collision_events, coll_rate}]
def vehicle_performance(d, c):
    rows = []
    if "vehicle" not in d.columns: return rows
    for veh, grp in d.groupby("vehicle"):
        n = len(grp)
        if n == 0: continue
        ok    = int(grp["flight_status"].str.startswith("Complete").sum())
        crash = int(grp["flight_status"].str.startswith("Collision").sum())
        batt  = int(grp["flight_status"].str.startswith("Incomplete").sum())
        canc  = int(grp["flight_status"].str.startswith("Cancelled").sum())
        bat_u = float(grp["bat_u"].mean()) if "bat_u" in grp.columns and not grp["bat_u"].isna().all() else 0
        bat_e = float(grp["bat_e"].mean()) if "bat_e" in grp.columns and not grp["bat_e"].isna().all() else 0
        da    = float(grp["da"].mean()) if "da" in grp.columns and not grp["da"].isna().all() else 0
        # Collision involvement
        veh_c = 0
        if not c.empty:
            veh_c = int((c.get("drone_a_vehicle")==veh).sum() + (c.get("drone_b_vehicle")==veh).sum())
        rows.append(dict(
            vehicle=str(veh), total=n,
            complete=ok, crash=crash, battery=batt, cancelled=canc,
            comp_pct=round(ok/n*100,1), crash_pct=round(crash/n*100,1),
            bat_avg=round(bat_u,1), bat_reserve=round(bat_e,1),
            avg_distance=round(da,1), collision_events=veh_c,
            coll_rate=round(veh_c/n*100,1),
        ))
    return sorted(rows, key=lambda x: -x["total"])

# layer_stats(d, c)
# → One row per altitude layer (1–4), used by the Safety page.
# Returns: list[{layer, label, total, complete, crash,
#                comp_pct, crash_pct, events, bat_avg}]
def layer_stats(d, c):
    rows = []
    for l in [1,2,3,4]:
        grp = d[d["layer"]==float(l)] if "layer" in d.columns else pd.DataFrame()
        n   = len(grp)
        if n == 0: continue
        ok    = int(grp["flight_status"].str.startswith("Complete").sum())
        crash = int(grp["flight_status"].str.startswith("Collision").sum())
        lname = str(grp["layer_name"].iloc[0]) if "layer_name" in grp.columns and not grp["layer_name"].isna().all() else f"L{l}"
        tc    = pd.DataFrame()
        if not c.empty:
            tc = c[c.get("drone_a_layer",pd.Series()).apply(lambda x: int(x)==l if pd.notna(x) else False)]
        rows.append(dict(
            layer=l, label=lname, total=n,
            complete=ok, crash=crash,
            comp_pct=round(ok/n*100,1), crash_pct=round(crash/n*100,1),
            events=len(tc),
            bat_avg=round(float(grp["bat_u"].mean()),1) if "bat_u" in grp.columns and not grp["bat_u"].isna().all() else 0,
        ))
    return rows

# collision_pairs(c)
# → Vehicle-pair type breakdown from the collision_type column
#   (e.g. "quad-hexa", "hexa-hexa").  Used by Overview and Safety pages.
# Returns: list[{pair, count, pct}]  top 15 pairs by frequency
def collision_pairs(c):
    if c.empty or "collision_type" not in c.columns: return []
    ct = c["collision_type"].value_counts().reset_index()
    ct.columns = ["pair","count"]
    total = len(c)
    return [{"pair":str(r["pair"]),"count":int(r["count"]),
             "pct":round(r["count"]/total*100,1)} for _,r in ct.head(15).iterrows()]

def path_run_battery_profile(d):
    if "path_run" not in d.columns or "bat_u" not in d.columns: return []
    rows = []
    for pr, grp in d.groupby("path_run"):
        vals = grp["bat_u"].dropna().values
        if len(vals) < 2: continue
        label = str(grp["path_label"].iloc[0]) if "path_label" in grp.columns else f"Run {pr}"
        rows.append(dict(
            path_run=int(pr), label=label,
            avg=round(float(np.mean(vals)),1), median=round(float(np.median(vals)),1),
            p25=round(float(np.percentile(vals,25)),1), p75=round(float(np.percentile(vals,75)),1),
            min=round(float(np.min(vals)),1), max=round(float(np.max(vals)),1),
            values=[_s(v) for v in vals[:300]],
        ))
    return sorted(rows, key=lambda x: x["path_run"])

def layer_crash_by_path(d):
    if "path_run" not in d.columns or "layer" not in d.columns: return []
    rows = []
    for pr, grp in d.groupby("path_run"):
        label = str(grp["path_label"].iloc[0]) if "path_label" in grp.columns else f"Run {pr}"
        for l in [1,2,3,4]:
            lg = grp[grp["layer"]==float(l)]
            n  = len(lg)
            if n == 0: continue
            crash = int(lg["flight_status"].str.startswith("Collision").sum())
            rows.append(dict(
                path_run=int(pr), label=label,
                layer=l, layer_name=f"L{l} · {(l-1)*50}m",
                total=n, crash=crash, crash_pct=round(crash/n*100,1),
            ))
    return rows

def vehicle_crash_by_layer(d):
    rows = []
    if "vehicle" not in d.columns or "layer" not in d.columns: return rows
    for veh, vgrp in d.groupby("vehicle"):
        for l in [1,2,3,4]:
            lgrp = vgrp[vgrp["layer"]==float(l)]
            n    = len(lgrp)
            if n == 0: continue
            crash = int(lgrp["flight_status"].str.startswith("Collision").sum())
            rows.append(dict(vehicle=str(veh), layer=l,
                             label=f"L{l} · {(l-1)*50}m",
                             total=n, crash=crash,
                             crash_pct=round(crash/n*100,1)))
    return rows

# algo_diagnostics(d, c)
# → All data for the Algorithm Diagnostics page.
# Six sub-sections assembled into a single dict:
#   "route_assignment"  → avg distance + crash rate per vehicle type
#   "separation_events" → every collision event with battery levels
#   "escalated_count"   → int — pairs that went Near Miss → Critical
#   "failed_pairs"      → list of escalated drone-pair events
#   "route_crash_dist"  → distance distributions: crashed vs safe drones
#   "risk_quadrant"     → 4-quadrant count from battery × distance matrix
#   "layer_fail"        → crash rate + avoidance failure rate per layer
def algo_diagnostics(d, c):
    """Data for Algorithm Diagnostics page."""
    out = {}

    # 1. Route assignment fairness: avg distance per vehicle
    veh_routes = []
    if "vehicle" in d.columns and "da" in d.columns:
        for veh, grp in d.groupby("vehicle"):
            vals = grp["da"].dropna().values
            if not len(vals): continue
            crash = int(grp["flight_status"].str.startswith("Collision").sum())
            n = len(grp)
            veh_routes.append(dict(
                vehicle=str(veh), n=n,
                avg_da=round(float(np.mean(vals)),1),
                median_da=round(float(np.median(vals)),1),
                p75_da=round(float(np.percentile(vals,75)),1),
                max_da=round(float(np.max(vals)),1),
                crash_pct=round(crash/n*100,1),
                values=[_s(v) for v in vals[:200]],
            ))
    out["route_assignment"] = veh_routes

    # 2. Separation margin: battery at time of collision
    sep_margins = []
    if not c.empty:
        for _, row in c.iterrows():
            sep_margins.append(dict(
                tick=_s(row.get("gtn")),
                type=str(row.get("type","—")),
                severity=str(row.get("severity","—")),
                bat_a=_s(row.get("bat_a")),
                bat_b=_s(row.get("bat_b")),
                veh_a=str(row.get("drone_a_vehicle","—")),
                veh_b=str(row.get("drone_b_vehicle","—")),
                layer=_s(row.get("drone_a_layer")),
                pair=str(row.get("collision_type","—")),
            ))
    out["separation_events"] = sep_margins

    # 3. Algorithm failure rate: how often Near Miss escalated to crash
    escalated = 0
    pairs_failed = []
    if not c.empty and "drone_a_id" in c.columns:
        import pandas as _pd
        c2 = c.copy()
        c2["pair"] = c2.apply(
            lambda r: f"D{min(int(r.drone_a_id),int(r.drone_b_id))}-D{max(int(r.drone_a_id),int(r.drone_b_id))}"
            if pd.notna(r.get("drone_a_id")) and pd.notna(r.get("drone_b_id")) else "?-?",
            axis=1)
        for pair, grp in c2.groupby("pair"):
            sevs = grp["severity"].tolist() if "severity" in grp.columns else []
            has_nm = any(s in ["Near Miss","Minor"] for s in sevs)
            has_crit = "Critical" in sevs
            if has_nm and has_crit:
                escalated += 1
                pairs_failed.append(dict(
                    pair=pair, n_events=len(grp),
                    first_tick=_s(grp["gtn"].min()) if "gtn" in grp.columns else None,
                    last_tick=_s(grp["gtn"].max()) if "gtn" in grp.columns else None,
                    outcome="Escalated to Critical",
                ))
    out["escalated_count"]  = escalated
    out["failed_pairs"]     = pairs_failed

    # 4. Crash predictability: route length distribution for crashed vs safe
    pred_data = {"crashed": [], "safe": []}
    if "da" in d.columns:
        crashed = d[d["flight_status"].str.startswith("Collision")]["da"].dropna().tolist()
        safe    = d[d["flight_status"].str.startswith("Complete")]["da"].dropna().tolist()
        pred_data["crashed"] = [_s(v) for v in crashed[:500]]
        pred_data["safe"]    = [_s(v) for v in safe[:500]]
    out["route_crash_dist"] = pred_data

    # 5. Battery-distance risk matrix: count drones in each risk quadrant
    risk_quad = {"high_risk":0,"vulnerable":0,"overloaded":0,"safe":0}
    if "bat_s" in d.columns and "da" in d.columns:
        da_med  = float(d["da"].median()) if not d["da"].isna().all() else 0
        bat_med = float(d["bat_s"].median()) if not d["bat_s"].isna().all() else 85
        for _, row in d.iterrows():
            da_v  = row.get("da")
            bat_v = row.get("bat_s")
            if pd.isna(da_v) or pd.isna(bat_v): continue
            if da_v > da_med and bat_v < bat_med:
                risk_quad["high_risk"] += 1
            elif da_v > da_med:
                risk_quad["overloaded"] += 1
            elif bat_v < bat_med:
                risk_quad["vulnerable"] += 1
            else:
                risk_quad["safe"] += 1
    out["risk_quadrant"] = risk_quad

    # 6. Per-layer algorithm failure rate (collisions / total events that could have been avoided)
    layer_fail = []
    for l in [1,2,3,4]:
        grp = d[d["layer"]==float(l)] if "layer" in d.columns else pd.DataFrame()
        n   = len(grp)
        if n == 0: continue
        crash = int(grp["flight_status"].str.startswith("Collision").sum())
        tc    = pd.DataFrame()
        if not c.empty and "drone_a_layer" in c.columns:
            tc = c[c["drone_a_layer"].apply(lambda x: int(x)==l if pd.notna(x) else False)]
        nm_count = int((tc["type"]=="Near Miss").sum()) if not tc.empty else 0
        dir_count= int((tc["type"]=="Direct").sum()) if not tc.empty else 0
        layer_fail.append(dict(
            layer=l, label=f"L{l} · {(l-1)*50}m",
            drones=n, crashes=crash, near_misses=nm_count,
            direct_hits=dir_count,
            crash_pct=round(crash/n*100,1),
            avoidance_fail_pct=round(dir_count/(nm_count+dir_count)*100,1)
                if (nm_count+dir_count)>0 else 0,
        ))
    out["layer_fail"] = layer_fail

    return out

# drone_positions(d, c)
# → Spatial data for the Airspace page Leaflet map.
# Returns:
#   "drones"    list[{id, x(lat), y(lon), layer, status, vehicle, bat_u}]
#   "conflicts" list[{x,y,bx,by, type, tick, drone_a, drone_b, veh_a, veh_b, layer}]
# x/y are WGS84 coordinates (cx/cy from enrich()).
def drone_positions(d, c):
    pos = d.dropna(subset=["cx","cy"])
    drones = [{"id":_s(r["drone_id"]),"x":_s(r["cx"]),"y":_s(r["cy"]),
               "layer":_s(r.get("layer")),"status":str(r["flight_status"]),
               "vehicle":str(r.get("vehicle","—")),"bat_u":_s(r.get("bat_u"))}
              for _, r in pos.iterrows()]
    conflicts = []
    if not c.empty:
        for _, r in c.iterrows():
            conflicts.append({"x":_s(r.get("drone_a_coord_x")),"y":_s(r.get("drone_a_coord_y")),
                               "bx":_s(r.get("drone_b_coord_x")),"by":_s(r.get("drone_b_coord_y")),
                               "type":str(r.get("type","—")),"tick":_s(r.get("gtn")),
                               "drone_a":_s(r.get("drone_a_id")),"drone_b":_s(r.get("drone_b_id")),
                               "veh_a":str(r.get("drone_a_vehicle","—")),
                               "veh_b":str(r.get("drone_b_vehicle","—")),
                               "layer":_s(r.get("drone_a_layer"))})
    return {"drones": drones, "conflicts": conflicts}

# rolling_rate(c)
# → Collision count per simulation tick plus a 5-tick rolling average.
# Used by the Temporal page rolling rate chart.
# Returns: {ticks:[int], raw:[int], rolling:[float]}
def rolling_rate(c):
    if c.empty or "gtn" not in c.columns: return {"ticks":[],"raw":[],"rolling":[]}
    ct = c.dropna(subset=["gtn"])
    if ct.empty: return {"ticks":[],"raw":[],"rolling":[]}
    gmin,gmax = int(ct["gtn"].min()), int(ct["gtn"].max())
    ticks = list(range(gmin,gmax+1))
    tc    = ct["gtn"].value_counts()
    raw   = [int(tc.get(t,0)) for t in ticks]
    roll  = pd.Series(raw).rolling(5,min_periods=1).mean().round(2).tolist()
    return {"ticks":ticks,"raw":raw,"rolling":roll}

# fleet_density_data(d, c)
# → Estimated airborne drone count per tick alongside collision counts.
# "airborne" is approximated: starts at total fleet size, decreases by 2
# for each collision event (each removes both drones involved).
# Used by the Temporal page airborne chart and collision-vs-airborne chart.
# Returns: {ticks:[int], airborne:[int], collision_counts:[int]}
def fleet_density_data(d, c):
    if c.empty or "gtn" not in c.columns: return {"ticks":[],"airborne":[],"collision_counts":[]}
    c_t = c.dropna(subset=["gtn"])
    if c_t.empty: return {"ticks":[],"airborne":[],"collision_counts":[]}
    gmin,gmax = int(c_t["gtn"].min()), int(c_t["gtn"].max())
    ticks = list(range(gmin,gmax+1))
    total = len(d)
    coll_counts = c_t["gtn"].value_counts()
    cum_lost = 0; airborne = []; coll_list = []
    for t in ticks:
        evts = int(coll_counts.get(t,0))
        coll_list.append(evts)
        cum_lost += evts*2
        airborne.append(max(0,total-cum_lost))
    return {"ticks":ticks,"airborne":airborne,"collision_counts":coll_list}

# bat_reserve_layer(d)
# → Every drone's battery-end value paired with its layer and status.
# Used by the Fleet page battery-reserve-per-layer chart.
# Returns: list[{layer, bat_e, flight_status}]
def bat_reserve_layer(d):
    if "layer" not in d.columns or "bat_e" not in d.columns: return []
    bd = d.dropna(subset=["layer","bat_e"])
    return [{"layer":_s(r["layer"]),"bat_e":_s(r["bat_e"]),
             "flight_status":str(r["flight_status"])} for _,r in bd.iterrows()]

# fleet_funnel(d)
# → Flight-outcome breakdown used by the Fleet page status funnel chart.
# Exact flight_status strings counted (not startswith — exact match):
#   "Collision — Node to Node"  "Collision — Proximity"
#   "Incomplete — Battery"  "Cancelled — In-flight"  "Cancelled — Pre-flight"
# Returns: {total, complete, coll_direct, coll_prox,
#            batt_fail, canc_inflight, canc_preflight}
def fleet_funnel(d):
    N = len(d)
    if N == 0: return {"total":0}
    return dict(
        total=N,
        complete=int(d["flight_status"].str.startswith("Complete").sum()),
        coll_direct=int((d["flight_status"]=="Collision — Node to Node").sum()),
        coll_prox=int((d["flight_status"]=="Collision — Proximity").sum()),
        batt_fail=int((d["flight_status"]=="Incomplete — Battery").sum()),
        canc_inflight=int((d["flight_status"]=="Cancelled — In-flight").sum()),
        canc_preflight=int((d["flight_status"]=="Cancelled — Pre-flight").sum()),
    )

# ml_risk(d)
# → Risk scoring for the Predictive page.
# Trains a RandomForestClassifier on [bat_s, layer, da] to predict
# "Collision" flight status.  Labels each drone with a risk_score (0–1)
# and risk_tier (Low / Medium / High).
# Also runs IsolationForest (contamination=0.15) to flag anomalies.
# Requires scikit-learn.  Returns an error dict if not installed.
# Feature thresholds: score ≥ 0.3 → Medium, ≥ 0.6 → High.
# Returns: {drones[], feature_importance{}, risk_tiers{},
#            total_analyzed, confusion{tp,fp,tn,fn}, n_anomalies}
def ml_risk(d):
    try:
        from sklearn.ensemble import RandomForestClassifier, IsolationForest
    except ImportError:
        return {"error":"scikit-learn not installed","drones":[],"feature_importance":{},"risk_tiers":{}}
    feats  = ["bat_s","layer","da"]
    df     = d.dropna(subset=feats+["flight_status"]).copy()
    if len(df) < 10: return {"error":"Insufficient data","drones":[],"feature_importance":{},"risk_tiers":{}}
    df["crashed"] = df["flight_status"].str.startswith("Collision").astype(int)
    X = df[feats].values; y = df["crashed"].values
    rf = RandomForestClassifier(n_estimators=100,random_state=42,class_weight="balanced")
    rf.fit(X,y); probs = rf.predict_proba(X)[:,1]
    df["risk_score"] = probs
    iso = IsolationForest(contamination=0.15,random_state=42)
    df["anomaly"] = (iso.fit_predict(X)==-1).astype(int)
    df["risk_tier"] = "Low"
    df.loc[probs>=0.3,"risk_tier"] = "Medium"
    df.loc[probs>=0.6,"risk_tier"] = "High"
    feat_labels = ["Battery at Launch","Altitude Layer","Distance Flown"]
    fi = {feat_labels[i]: round(float(rf.feature_importances_[i]),3) for i in range(len(feat_labels))}
    tiers = df["risk_tier"].value_counts().to_dict()
    tp=int(((probs>=0.5)&(y==1)).sum()); fp=int(((probs>=0.5)&(y==0)).sum())
    tn=int(((probs<0.5)&(y==0)).sum());  fn=int(((probs<0.5)&(y==1)).sum())
    drones_out = []
    for _,row in df.iterrows():
        drones_out.append({"id":_s(row["drone_id"]),"x":_s(row.get("cx")),"y":_s(row.get("cy")),
                           "layer":_s(row["layer"]),"risk_score":round(float(row["risk_score"]),3),
                           "risk_tier":row["risk_tier"],"anomaly":int(row["anomaly"]),
                           "status":str(row["flight_status"]),"bat_s":_s(row["bat_s"]),
                           "da":_s(row["da"]),"crashed":int(row["crashed"])})
    return {"drones":drones_out,"feature_importance":fi,
            "risk_tiers":{"Low":int(tiers.get("Low",0)),"Medium":int(tiers.get("Medium",0)),"High":int(tiers.get("High",0))},
            "total_analyzed":len(df),"confusion":{"tp":tp,"fp":fp,"tn":tn,"fn":fn},
            "n_anomalies":int(df["anomaly"].sum())}

# vehicle_radar_data(d, c)
# → 5-axis radar chart data per vehicle type for the Predictive page.
# Axes (all normalised 0–100, higher = better):
#   completion   → % drones with "Complete" status
#   bat_endurance → 100 − battery_failure_rate%
#   energy_eff   → 100 − avg_battery_consumed%
#   reserve      → avg battery remaining at landing%
#   reliability  → 100 − cancellation_rate%
# Returns: {vehicles:[{vehicle, completion, bat_endurance, energy_eff,
#                       reserve, reliability}],
#            axes:[str x 5]}
def vehicle_radar_data(d, c):
    rows = []
    for veh, grp in d.groupby("vehicle"):
        n = len(grp)
        if n == 0: continue
        comp      = round(grp["flight_status"].str.startswith("Complete").sum()/n*100,1)
        bat_fail  = round((grp["flight_status"]=="Incomplete — Battery").sum()/n*100,1)
        bat_u     = round(float(grp["bat_u"].mean()),1) if "bat_u" in grp.columns and not grp["bat_u"].isna().all() else 50
        bat_e     = round(float(grp["bat_e"].mean()),1) if "bat_e" in grp.columns and not grp["bat_e"].isna().all() else 50
        canc      = round(grp["flight_status"].str.startswith("Cancelled").sum()/n*100,1)
        rows.append({"vehicle":veh,"completion":comp,"bat_endurance":round(100-bat_fail,1),
                     "energy_eff":round(100-bat_u,1),"reserve":bat_e,
                     "reliability":round(100-canc,1)})
    return {"vehicles":rows,"axes":["Completion","Battery Endurance","Energy Efficiency","Reserve at Landing","Launch Reliability"]}

# conflict_escalation(c)
# → Groups collision events by drone pair and identifies pairs that
# escalated from Near Miss / Minor to Critical within the same trial.
# Pairs are sorted: escalated pairs first, then by first_tick.
# Used by the Predictive page escalation list.
# Returns: {pairs:[{pair, events[], n_events, escalated, final,
#                   first_tick, last_tick}],
#            ticks:[int]}
def conflict_escalation(c):
    if c.empty: return {"pairs":[],"ticks":[]}
    if "drone_a_id" not in c.columns: return {"pairs":[],"ticks":[]}
    c2 = c.copy()
    c2["pair"] = c2.apply(
        lambda r: f"D{min(int(r.drone_a_id),int(r.drone_b_id))}-D{max(int(r.drone_a_id),int(r.drone_b_id))}"
        if pd.notna(r.get("drone_a_id")) and pd.notna(r.get("drone_b_id")) else "?-?",
        axis=1)
    pairs_out = []
    for pair, grp in c2.groupby("pair"):
        grp = grp.sort_values("gtn")
        events = [{"tick":int(row["gtn"]),"type":str(row["type"]),"severity":str(row["severity"])}
                  for _,row in grp.iterrows() if "gtn" in grp.columns]
        sevs = [e["severity"] for e in events]
        escalated = ("Near Miss" in sevs or "Minor" in sevs) and "Critical" in sevs
        pairs_out.append({"pair":pair,"events":events,"n_events":len(events),
                          "escalated":escalated,"final":sevs[-1] if sevs else "—",
                          "first_tick":events[0]["tick"] if events else 0,
                          "last_tick":events[-1]["tick"] if events else 0})
    pairs_out.sort(key=lambda x: (not x["escalated"],x["first_tick"]))
    all_ticks = sorted(c2["gtn"].dropna().unique().tolist())
    return {"pairs":pairs_out,"ticks":[int(t) for t in all_ticks]}

# ═══════════════════════════════════════════════════════════════════════════════
# API ROUTES
# One FastAPI route per frontend page or action.
# Every route (except connect) calls _get(tid, path_runs) to get enriched
# DataFrames, then calls one or more data-builder functions, and returns
# the result via J() as a JSONResponse.
# The frontend api.js file mirrors these routes exactly.
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/")
async def root():
    return FileResponse(str(FRONTEND / "index.html"))

@app.post("/api/connect/postgres")
async def connect_pg(
    host: str = Form("localhost"), port: int = Form(5432),
    dbname: str = Form("utm_geo_db"), user: str = Form("postgres"),
    password: str = Form("omkar7781"),
):
    try:
        data = load_postgres(host, port, dbname, user, password)
        _S["drones"]     = data["drones"]
        _S["collisions"] = data["collisions"]
        _S["source"]     = "postgresql"
        _S["trials"]     = sorted(_S["drones"]["trial_id"].unique().tolist())
        return {"ok": True, "trials": _S["trials"], "source": "postgresql"}
    except Exception as e:
        raise HTTPException(400, str(e))

@app.post("/api/connect/excel")
async def connect_excel(file: UploadFile = File(...)):
    try:
        buf = io.BytesIO(await file.read())
        buf.name = file.filename
        data = load_excel(buf)
        _S["drones"]     = data["drones"]
        _S["collisions"] = data["collisions"]
        _S["source"]     = "excel"
        _S["trials"]     = sorted(_S["drones"]["trial_id"].unique().tolist())
        return {"ok": True, "trials": _S["trials"], "source": "excel"}
    except Exception as e:
        raise HTTPException(400, str(e))

@app.get("/api/trials")
async def get_trials():
    return {"trials": _S["trials"], "source": _S["source"]}

@app.get("/api/trial/{tid}/path_runs")
async def get_path_runs(tid: str):
    if _S["drones"].empty: return {"path_runs":[]}
    d = _S["drones"][_S["drones"]["trial_id"]==tid]
    prs = sorted(d["path_run"].dropna().unique().tolist())
    labels = {}
    for pr in prs:
        grp = d[d["path_run"]==pr]
        labels[int(pr)] = str(grp["path_label"].iloc[0]) if "path_label" in grp.columns and not grp["path_label"].isna().all() else f"Run {pr}"
    return {"path_runs":[int(p) for p in prs],"labels":labels}

@app.get("/api/trial/{tid}/kpis")
async def get_kpis(tid: str, path_runs: Optional[str]=None):
    d,c = _get(tid, path_runs)
    return J(compute_kpis(d,c))

@app.get("/api/trial/{tid}/overview")
async def get_overview(tid: str, path_runs: Optional[str]=None):
    d,c  = _get(tid, path_runs)
    k    = compute_kpis(d,c)
    veh  = {str(v):int(n) for v,n in d["vehicle"].value_counts().items()} if "vehicle" in d.columns else {}
    lyr  = {str(l):int((d["layer"]==l).sum()) for l in [1,2,3,4]} if "layer" in d.columns else {}
    coll_log = []
    if not c.empty:
        for _,row in c.head(50).iterrows():
            lat = row.get("drone_a_coord_x")
            lon = row.get("drone_a_coord_y")
            coll_log.append({"path_run":_s(row.get("path_run")),"tick":_s(row.get("gtn")),
                             "type":str(row.get("type","—")),"severity":str(row.get("severity","—")),
                             "drone_a":_s(row.get("drone_a_id")),"drone_b":_s(row.get("drone_b_id")),
                             "veh_a":str(row.get("drone_a_vehicle","—")),
                             "veh_b":str(row.get("drone_b_vehicle","—")),
                             "layer":_s(row.get("drone_a_layer")),
                             "pair":str(row.get("collision_type","—")),
                             "lat": float(lat) if lat is not None and str(lat) not in ("nan","None","—") else None,
                             "lon": float(lon) if lon is not None and str(lon) not in ("nan","None","—") else None})
    outcomes = d["flight_status"].value_counts().reset_index()
    outcomes.columns = ["status","count"]
    return J({"kpis":k,"outcomes":{"statuses":outcomes["status"].tolist(),"counts":outcomes["count"].tolist()},
              "vehicles":veh,"layers":lyr,"coll_log":coll_log,
              "path_run_comparison":path_run_comparison(d,c),
              "collision_pairs":collision_pairs(c)})

@app.get("/api/trial/{tid}/spatial")
async def get_spatial(tid: str, path_runs: Optional[str]=None):
    d,c = _get(tid, path_runs)
    return J(drone_positions(d,c))

@app.get("/api/trial/{tid}/safety")
async def get_safety(tid: str, path_runs: Optional[str]=None):
    d,c = _get(tid, path_runs)
    k   = compute_kpis(d,c)
    sev = c["severity"].value_counts().reset_index() if not c.empty else pd.DataFrame(columns=["severity","count"])
    sev.columns = ["severity","count"]
    veh_hits = []
    if not c.empty and "drone_a_vehicle" in c.columns:
        vh = pd.concat([c["drone_a_vehicle"],c["drone_b_vehicle"]]).value_counts().reset_index()
        vh.columns = ["vehicle","collisions"]
        fc = d["vehicle"].value_counts().to_dict() if "vehicle" in d.columns else {}
        vh["fleet"]    = vh["vehicle"].map(fc).fillna(0)
        vh["collision_rate"] = (vh["collisions"]/vh["fleet"].replace(0,np.nan)*100).round(1)
        vh["hits"] = vh["collisions"]
        vh["fleet_count"] = vh["fleet"].astype(int)
        veh_hits = vh[["vehicle","hits","fleet_count","collision_rate"]].to_dict("records")
    bat_crash = pd.concat([c["bat_a"],c["bat_b"]]).dropna().values if not c.empty and "bat_a" in c.columns else np.array([])
    bat_kde = {"x":[],"y":[]}
    if len(bat_crash)>=3:
        try:
            from scipy.stats import gaussian_kde
            kde = gaussian_kde(bat_crash,bw_method=0.4)
            xs  = np.linspace(0,100,200)
            bat_kde = {"x":xs.tolist(),"y":kde(xs).tolist(),"obs":bat_crash.tolist()}
        except: pass
    return J({"kpis":k,"severity":{"statuses":sev["severity"].tolist(),"counts":sev["count"].tolist()},
              "vehicle_hits":veh_hits,"bat_kde":bat_kde,
              "layer_stats":layer_stats(d,c),
              "vehicle_perf":vehicle_performance(d,c),
              "collision_pairs":collision_pairs(c),
              "vehicle_crash_layer":vehicle_crash_by_layer(d),
              "escalation":conflict_escalation(c)})

@app.get("/api/trial/{tid}/fleet")
async def get_fleet(tid: str, path_runs: Optional[str]=None):
    d,c  = _get(tid, path_runs)
    k    = compute_kpis(d,c)
    bk   = {}
    for field,key in [("bat_s","start"),("bat_u","consumed"),("bat_e","end")]:
        vals = d[field].dropna().values if field in d.columns else np.array([])
        if len(vals)>=3:
            try:
                from scipy.stats import gaussian_kde
                kde = gaussian_kde(vals,bw_method=0.3)
                xs  = np.linspace(max(0,vals.min()-2),min(100,vals.max()+2),200)
                bk[key] = {"x":xs.tolist(),"y":kde(xs).tolist()}
            except: bk[key] = {"x":[],"y":[]}
        else: bk[key] = {"x":[],"y":[]}
    status_kdes = []
    if "bat_u" in d.columns and "flight_status" in d.columns:
        from scipy.stats import gaussian_kde
        for status,grp in d.groupby("flight_status"):
            vals = grp["bat_u"].dropna().values
            if len(vals)<2: continue
            try:
                kde = gaussian_kde(vals,bw_method=0.4)
                xs  = np.linspace(max(0,vals.min()-2),vals.max()+2,200)
                status_kdes.append({"status":status,"x":xs.tolist(),"y":kde(xs).tolist()})
            except: pass
    drain_pts = []
    if "da" in d.columns and "bat_u" in d.columns:
        bd = d.dropna(subset=["da","bat_u","vehicle"])
        bd = bd[bd["da"]>0]
        drain_pts = [{"da":_s(r["da"]),"bat_u":_s(r["bat_u"]),"vehicle":str(r["vehicle"]),"status":str(r["flight_status"])} for _,r in bd.iterrows()]
    return J({"kpis":k,"bat_kde":bk,"bat_kde_by_status":status_kdes,
              "fleet_funnel":fleet_funnel(d),
              "bat_drain_points":drain_pts,
              "bat_reserve_layer":bat_reserve_layer(d),
              "vehicle_perf":vehicle_performance(d,c),
              "path_run_battery":path_run_battery_profile(d),
              "layer_crash_trend":layer_crash_by_path(d),
              "vehicle_crash_layer":vehicle_crash_by_layer(d)})

@app.get("/api/trial/{tid}/temporal")
async def get_temporal(tid: str, path_runs: Optional[str]=None):
    d,c = _get(tid, path_runs)
    dur_by_status = []
    if "duration" in d.columns and "flight_status" in d.columns:
        for status,grp in d.groupby("flight_status"):
            vals = grp["duration"].dropna().values
            if len(vals)>=2:
                dur_by_status.append({"status":status,"values":[_s(v) for v in vals[:300]]})
    return J({"rolling_rate":rolling_rate(c),
              "fleet_density":fleet_density_data(d,c),
              "dur_by_status":dur_by_status,
              "coll_by_tick":c[["gtn","type","severity","drone_a_id","drone_b_id",
                                 "drone_a_vehicle","drone_b_vehicle","drone_a_layer"]].dropna(subset=["gtn"]).to_dict("records") if not c.empty else []})

@app.get("/api/trial/{tid}/ml_risk")
async def get_ml_risk(tid: str, path_runs: Optional[str]=None):
    d,c = _get(tid, path_runs)
    return J({"ml":ml_risk(d), "vehicle_radar":vehicle_radar_data(d,c),
              "conflict_escalation":conflict_escalation(c)})

@app.get("/api/trial/{tid}/algo_diag")
async def get_algo_diag(tid: str, path_runs: Optional[str]=None):
    d,c = _get(tid, path_runs)
    return J(algo_diagnostics(d,c))

@app.get("/api/multitrail_intel")
async def get_multitrail_intel(trial: Optional[str]=None, path_runs: Optional[str]=None):
    """
    Cross-trial intelligence — one summary row per trial (or per path run
    if a specific trial + path_runs are passed).
    When called without args: iterates ALL trial_ids in _S.
    When called with ?trial=X&path_runs=1,2: groups by path_run within X.
    Used by the Intel page.
    Returns: {trials:[{trial, fleet, complete, comp_pct, collisions,
                        coll_rate, bat_used, efficiency, high_risk_drones}],
              n_trials: int}
    """
    if _S["drones"].empty:
        return J({"trials": [], "n_trials": 0})
    d_all = _S["drones"]
    c_all = _S["collisions"] if not _S["collisions"].empty else pd.DataFrame()
    # If a specific trial is selected, filter to just that trial (and its path runs)
    if trial and "trial_id" in d_all.columns:
        d_all = d_all[d_all["trial_id"] == trial]
        if not c_all.empty and "trial_id" in c_all.columns:
            c_all = c_all[c_all["trial_id"] == trial]
        if path_runs and "path_run" in d_all.columns:
            prs = [p.strip() for p in path_runs.split(",") if p.strip()]
            if prs:
                d_all = d_all[d_all["path_run"].isin(prs)]
                if not c_all.empty and "path_run" in c_all.columns:
                    c_all = c_all[c_all["path_run"].isin(prs)]
    rows = []
    group_col = "path_run" if (path_runs and "path_run" in d_all.columns) else "trial_id"
    for tid, grp in d_all.groupby(group_col):
        n = len(grp)
        if n == 0: continue
        tc = c_all[c_all.get(group_col, pd.Series(dtype=str)) == tid] if not c_all.empty and group_col in c_all.columns else pd.DataFrame()
        complete = int(grp["flight_status"].str.startswith("Complete").sum())
        collisions = len(tc)
        comp_pct = round(complete / n * 100, 1)
        coll_rate = round(collisions / n * 100, 1)
        bat_used = round(float(grp["bat_u"].mean()), 1) if "bat_u" in grp.columns and not grp["bat_u"].isna().all() else 0.0
        da_avg = round(float(grp["da"].mean()), 1) if "da" in grp.columns and not grp["da"].isna().all() else 0.0
        efficiency = round(da_avg / max(bat_used, 1), 2) if bat_used > 0 else 0.0
        high_risk = int(grp["flight_status"].str.startswith("Collision").sum())
        rows.append(dict(
            trial=str(tid), fleet=n,
            complete=complete, comp_pct=comp_pct,
            collisions=collisions, coll_rate=coll_rate,
            bat_used=bat_used, efficiency=efficiency,
            high_risk_drones=high_risk,
        ))
    rows.sort(key=lambda x: x["trial"])
    return J({"trials": rows, "n_trials": len(rows)})

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
